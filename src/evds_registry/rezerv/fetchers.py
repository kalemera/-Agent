"""Veri kaynağı fetcher'ları — EVDS API, URDL ZIP, Taraflı Swap PDF.

Üç fetcher:
    fetch_evds_series()       — EVDS API üzerinden 51 seri (chunk'lı)
    fetch_urdl_zip()          — TCMB sayfası → Haftalık ZIP → Excel parse
    fetch_tarafli_swap_pdf()  — TCMB Piyasa Verileri → PDF → tablo parse
"""

from __future__ import annotations

import datetime as _dt
import io
import re
import warnings
import zipfile
from typing import Any
from urllib.parse import urlparse

import pandas as pd
import requests

from .config import (
    ALL_TICKERS,
    PDF_KEYWORDS,
    PIYASA_VERILERI_URL,
    URDL_PAGE_URL,
    URDL_ZIP_TITLE_PREFIX,
)


_DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Güvenlik sabitleri — security audit (2026-05-07) önerileri
_ALLOWED_HOST_SUFFIXES: tuple[str, ...] = ("tcmb.gov.tr",)
"""SSRF koruması: sadece bu host suffix'lerine HTTP istek atılır."""

_MAX_ZIP_BYTES: int = 50 * 1024 * 1024            # 50 MB
_MAX_EXCEL_BYTES: int = 100 * 1024 * 1024          # 100 MB (ZIP içi)
_MAX_PDF_BYTES: int = 50 * 1024 * 1024             # 50 MB


def _assert_allowed_host(url: str) -> None:
    """SSRF koruması: hostname tcmb.gov.tr ile bitmiyorsa reddet.

    Hem absolute (https://www.tcmb.gov.tr/...) hem ileride relatif birleştirilen
    URL'ler için son kontrol noktası.
    """
    host = urlparse(url).hostname or ""
    host_lower = host.lower()
    if not any(host_lower == s or host_lower.endswith("." + s) or host_lower == s
               for s in _ALLOWED_HOST_SUFFIXES):
        raise ValueError(
            f"İzin verilmeyen host: {host!r}. Sadece {_ALLOWED_HOST_SUFFIXES} izinli."
        )


def _safe_get(url: str, headers: dict | None = None, timeout: int = 30,
              max_bytes: int | None = None) -> bytes:
    """Güvenli HTTP GET — host whitelist + boyut kapısı + redirect limiti.

    Args:
        url: İstek URL'i (tcmb.gov.tr suffix gerekli).
        headers: HTTP headers.
        timeout: Bağlantı timeout (saniye).
        max_bytes: Tam body boyutu üst sınırı. Aşılırsa ValueError.
    """
    _assert_allowed_host(url)
    sess = requests.Session()
    sess.max_redirects = 3
    resp = sess.get(url, headers=headers, timeout=timeout, stream=True,
                    allow_redirects=True)
    resp.raise_for_status()

    # Redirect olduysa final URL host kontrolü
    if resp.url != url:
        _assert_allowed_host(resp.url)

    chunks = []
    total = 0
    for chunk in resp.iter_content(chunk_size=64 * 1024):
        if not chunk:
            continue
        total += len(chunk)
        if max_bytes is not None and total > max_bytes:
            resp.close()
            raise ValueError(
                f"Yanıt boyutu {max_bytes} byte sınırını aştı (URL: {url})."
            )
        chunks.append(chunk)
    resp.close()
    return b"".join(chunks)


# =============================================================================
# 1) EVDS API Fetcher
# =============================================================================

def _normalize_ticker_for_evds_lib(ticker: str) -> str:
    """EVDS lib bazen 'TP.AB.A02' yerine 'TP_AB_A02' bekleyebilir.
    Resmi `evdsAPI.get_data()` original noktayı destekliyor — değiştirme.
    """
    return ticker


def fetch_evds_series(
    tickers: list[str] | None = None,
    start_date: str = "15-01-2013",
    end_date: str = "",
    api_key: str | None = None,
    frequency: str = "",
    chunk_size: int = 25,
) -> pd.DataFrame:
    """EVDS API'den seri verilerini çek.

    Args:
        tickers: Çekilecek tickerlar. None ise ALL_TICKERS kullanılır.
        start_date: 'dd-mm-yyyy' formatında başlangıç tarihi.
        end_date: Boş bırakılırsa bugün.
        api_key: None ise EVDS_API_KEY env var'dan alınır.
        frequency: '5'=aylık, '3'=haftalık, '1'=günlük (boş = default).
        chunk_size: Tek istekte gönderilecek max ticker sayısı.

    Returns:
        DataFrame[index=Tarih datetime, columns=ticker_kısa_ad]
        Sütun adları orijinal ticker'dan kısaltılır:
            'TP_AB_A02' → 'AB.A02'
            'TP_BL001'  → 'BL001'
            'TP_DK_USD_A_YTL' → 'DK.USD.A.YTL'
            'TP_REZVARPD_K1'  → 'REZVARPD.K1'
            'TP_DOVVARNC_K18' → 'DOVVARNC.K18'
    """
    from evds import evdsAPI

    if tickers is None:
        tickers = list(ALL_TICKERS)

    if api_key is None:
        import os
        api_key = (
            os.getenv("EVDS_API_KEY")
            or os.getenv("EVDS_KEY")
            or ""
        )
    if not api_key:
        raise ValueError("EVDS API key gereklidir (EVDS_API_KEY env var veya api_key parametresi)")

    if not end_date:
        end_date = _dt.date.today().strftime("%d-%m-%Y")

    client = evdsAPI(api_key)

    # Chunk'lı çekme (URL uzunluğu ve rate limit korunur)
    chunks: list[pd.DataFrame] = []
    for i in range(0, len(tickers), chunk_size):
        chunk = tickers[i : i + chunk_size]
        try:
            kwargs: dict[str, Any] = {}
            if frequency:
                kwargs["frequency"] = frequency
            df = client.get_data(chunk, startdate=start_date, enddate=end_date, **kwargs)
        except Exception as e:
            warnings.warn(f"EVDS chunk {i}-{i+chunk_size} fail: {e}", stacklevel=2)
            continue
        if df is None or df.empty:
            continue
        chunks.append(df)

    if not chunks:
        return pd.DataFrame()

    # Tüm chunk'ları Tarih üzerinden join
    merged = chunks[0]
    for df in chunks[1:]:
        if "Tarih" in merged.columns and "Tarih" in df.columns:
            merged = merged.merge(df, on="Tarih", how="outer")
        else:
            # Index bazlı concat (fallback)
            merged = pd.concat([merged, df], axis=1)

    # Tarih kolonu → DatetimeIndex
    merged = _normalize_date_column(merged)

    # Sütun adlarını kısalt (TP_ prefix kaldır, _ → .)
    merged = _shorten_evds_column_names(merged)

    return merged


def _normalize_date_column(df: pd.DataFrame) -> pd.DataFrame:
    """Tarih sütununu DatetimeIndex'e çevir."""
    out = df.copy()
    date_col = None
    for cand in ["Tarih", "DATE", "tarih", "date"]:
        if cand in out.columns:
            date_col = cand
            break
    if date_col is None:
        return out

    # EVDS lib bazen "01-01-2024" formatında string döner
    parsed = pd.to_datetime(out[date_col], format="%d-%m-%Y", errors="coerce")
    if parsed.isna().all():
        parsed = pd.to_datetime(out[date_col], errors="coerce")
    out[date_col] = parsed
    out = out.dropna(subset=[date_col]).set_index(date_col).sort_index()
    out.index.name = "Tarih"

    # UNIXTIME varsa kaldır
    out = out.drop(columns=["UNIXTIME", "YEARWEEK"], errors="ignore")
    return out


def _shorten_evds_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """'TP_AB_A02' → 'AB.A02' gibi kısaltma."""
    rename = {}
    for col in df.columns:
        if not isinstance(col, str):
            continue
        if col.startswith("TP_"):
            short = col[3:].replace("_", ".")
            rename[col] = short
    return df.rename(columns=rename)


# =============================================================================
# 2) URDL ZIP Fetcher
# =============================================================================

def fetch_urdl_zip(
    url: str = URDL_PAGE_URL,
    timeout: int = 30,
) -> pd.DataFrame:
    """TCMB URDL Haftalık ZIP scrape'i.

    Adımlar (M kod 1-451):
        1. Sayfada 'Haftalık Uluslararası Rezervler' başlıklı ZIP linkini bul
        2. ZIP'i indir + magic check
        3. ZIP içindeki Excel'i parse et (openpyxl)
        4. Tarih sütunlu satırı başlık olarak tespit et
        5. Milyon USD → Milyar USD (÷1000)

    Returns:
        DataFrame[index=date, columns=URDL kalem adları]
        İlk sütun 'Kalem' kategori indexlenir, sonraki sütunlar tarihlerdir.
        Wide formatta döner — caller pivot edebilir.
    """
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook

    headers = {"User-Agent": _DEFAULT_USER_AGENT}

    # 1) Sayfa HTML
    page_bytes = _safe_get(url, headers=headers, timeout=timeout,
                           max_bytes=10 * 1024 * 1024)  # 10MB sayfa için fazla
    html_text = page_bytes.decode("utf-8", errors="replace")

    # 2) ZIP linki bul: title 'Haftalık Uluslararası Rezervler...' VE href .zip
    # NOT: TCMB sayfasında aynı başlıkla hem PDF hem ZIP linki var. Sadece
    # title eşleşmesi yetmez — `.zip` uzantı kontrolü zorunlu, aksi halde PDF
    # alınır ve ZIP magic check fail olur.
    soup = BeautifulSoup(html_text, "html.parser")
    zip_link = None
    for a in soup.find_all("a"):
        title = (a.get("title") or "").strip()
        href = (a.get("href") or "").strip()
        text = (a.get_text() or "").strip()
        if not href:
            continue
        title_match = title.lower().startswith(URDL_ZIP_TITLE_PREFIX.lower())
        text_zip = "ZIP" in text.upper()
        href_zip = ".zip" in href.lower()
        # Title VE (href .zip içerir VEYA text 'ZIP' içerir)
        if title_match and (href_zip or text_zip):
            zip_link = href
            break

    if not zip_link:
        raise RuntimeError(
            f"URDL ZIP linki bulunamadı (başlık: '{URDL_ZIP_TITLE_PREFIX}')"
        )

    if not zip_link.lower().startswith("http"):
        zip_link = "https://www.tcmb.gov.tr" + zip_link

    # SSRF doğrulama: relatif birleştirme veya absolute href kontrolü
    _assert_allowed_host(zip_link)

    # 3) ZIP indir (boyut kapısı + stream)
    zip_bytes = _safe_get(zip_link, headers=headers, timeout=timeout,
                          max_bytes=_MAX_ZIP_BYTES)
    if not zip_bytes.startswith(b"PK"):
        raise RuntimeError("İndirilen dosya ZIP değil (magic check fail)")

    # 4) ZIP içinden Excel oku — uncompressed boyut kontrolü
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    excel_name = None
    for name in zf.namelist():
        if name.lower().endswith((".xlsx", ".xls")):
            info = zf.getinfo(name)
            if info.file_size > _MAX_EXCEL_BYTES:
                raise ValueError(
                    f"ZIP içindeki Excel boyutu {info.file_size} byte sınırını aştı."
                )
            excel_name = name
            break
    if not excel_name:
        raise RuntimeError("ZIP içinde Excel dosyası bulunamadı")

    excel_bytes = zf.read(excel_name)

    # 5) Workbook'u oku — header satırını dinamik tespit etmek için raw rows
    # (pd.read_excel doğrudan header=N istemiyor; önce raw oku, header tespit
    # et, sonra DataFrame oluştur)
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    raw_rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        # Tuple → list, None olmayan en az 1 hücre varsa kabul et
        row_list = list(row)
        if any(c is not None for c in row_list):
            raw_rows.append(row_list)

    if not raw_rows:
        raise RuntimeError("URDL ZIP içindeki Excel boş.")

    # 6) Başlık satırı = en çok tarih hücresi içeren satır
    def _count_dates(row: list) -> int:
        return sum(1 for v in row if isinstance(v, (_dt.date, _dt.datetime)))

    header_idx = max(range(len(raw_rows)), key=lambda i: _count_dates(raw_rows[i]))
    headers_raw = raw_rows[header_idx]
    data_rows = raw_rows[header_idx + 1 :]

    # 7) Sütun adlarını string'e çevir (None → boş, date → ISO)
    column_names: list[str] = []
    seen: dict[str, int] = {}
    for idx, h in enumerate(headers_raw):
        if isinstance(h, (_dt.date, _dt.datetime)):
            base = pd.Timestamp(h).strftime("%Y-%m-%d")
        elif h is None:
            base = f"col_{idx}"
        else:
            base = str(h).strip() or f"col_{idx}"
        # Aynı isim varsa suffix ekle
        if base in seen:
            seen[base] += 1
            base = f"{base}__{seen[base]}"
        else:
            seen[base] = 0
        column_names.append(base)

    # 8) Data rows'u uniform uzunluğa getir (truncate/pad)
    ncols = len(column_names)
    normalized_data: list[list] = []
    for row in data_rows:
        if len(row) >= ncols:
            normalized_data.append(list(row[:ncols]))
        else:
            normalized_data.append(list(row) + [None] * (ncols - len(row)))

    df = pd.DataFrame(normalized_data, columns=column_names)

    # 9) URDL Excel formatında ilk sütun(lar) genelde boş — kaldır.
    # Boş kolonu tespit: tüm değerleri None/NaN olan sütunlar + header'ı boş
    # olan veya '' olan sütunlar
    def _is_empty_col(name: str) -> bool:
        ser = df[name]
        return ser.isna().all() or str(name).startswith("col_") and ser.isna().all()

    cols_to_drop = [c for c in df.columns if df[c].isna().all()]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)

    # 10) İlk kalan sütunu "Kalem" olarak adlandır
    if df.shape[1] == 0:
        raise RuntimeError("URDL ZIP Excel'inde geçerli sütun bulunamadı.")
    first_col = df.columns[0]
    if first_col != "Kalem":
        df = df.rename(columns={first_col: "Kalem"})

    # 10) Tarih-string sütunları pd.Timestamp'e çevir (header zaten tarih
    # olarak okunduysa "YYYY-MM-DD" formatında string)
    new_cols: dict[str, pd.Timestamp | str] = {}
    for col in df.columns:
        if col == "Kalem":
            continue
        # Yalnızca "YYYY-MM-DD" pattern'i Timestamp'e çevrilir
        try:
            ts = pd.to_datetime(col, format="%Y-%m-%d", errors="raise")
            new_cols[col] = ts
        except (ValueError, TypeError):
            # Tarih değil — string olarak bırak (örn rename collision suffix'leri)
            pass
    if new_cols:
        df = df.rename(columns=new_cols)

    # 11) Sayısal sütunlar (tarih kolonları) Milyon USD → Milyar USD (÷1000)
    for col in df.columns:
        if col == "Kalem":
            continue
        # Sütun değerleri Excel'den int/float/None olarak gelmiş — pd.to_numeric
        # ile pandas Series'e zorla, dtype coerce et.
        numeric_series = pd.to_numeric(df[col], errors="coerce")
        df[col] = numeric_series / 1000.0

    # 12) Boş Kalem satırlarını sil
    df = df[df["Kalem"].notna()].reset_index(drop=True)

    return df


def urdl_extract_kalem_series(
    urdl_df: pd.DataFrame,
    keywords: list[str],
    output_name: str = "value",
) -> pd.DataFrame:
    """URDL wide formatından bir kalemi tarihli long formata çevir.

    Args:
        urdl_df: fetch_urdl_zip() çıktısı
        keywords: Kalem sütununda hepsi bulunmalı (case-insensitive, TR-normalize)
        output_name: Çıktı sütun adı

    Returns:
        DataFrame[index=Tarih, columns={output_name}]
    """
    def _normalize_tr(s: str) -> str:
        s = (s or "").lower()
        # Turkish capital İ lowercases to i + U+0307 (combining dot); strip it.
        s = s.replace("̇", "")
        for tr, en in [("ı", "i"), ("ş", "s"), ("ğ", "g"), ("ç", "c"), ("ö", "o"), ("ü", "u")]:
            s = s.replace(tr, en)
        return s

    keywords_n = [_normalize_tr(k) for k in keywords]

    def _matches(kalem: str) -> bool:
        n = _normalize_tr(str(kalem))
        return all(k in n for k in keywords_n)

    matches = urdl_df[urdl_df["Kalem"].apply(_matches)]
    if matches.empty:
        return pd.DataFrame(columns=[output_name])

    row = matches.iloc[0]
    date_cols = [c for c in urdl_df.columns if c != "Kalem"]
    series = row[date_cols]
    series.index = pd.to_datetime(series.index, errors="coerce")
    series = series[series.index.notna()]

    out = pd.DataFrame({output_name: series.values}, index=series.index)
    out.index.name = "Tarih"
    return out


# =============================================================================
# 3) Taraflı Swap PDF Fetcher
# =============================================================================

def fetch_tarafli_swap_pdf(
    url: str = PIYASA_VERILERI_URL,
    timeout: int = 30,
) -> pd.DataFrame:
    """TCMB Piyasa Verileri sayfasından günlük PDF scrape'i.

    Adımlar (M kod 2127-2580):
        1. Piyasa Verileri sayfasında 'tcmb+taraf+swap+islem' kelimeleri
           geçen PDF linkini bul
        2. PDF'i indir + magic check (%PDF)
        3. pdfplumber ile tabloları çıkar
        4. Başlık satırı (Valör/Value) tespit + sütun temizleme
        5. Milyon USD → Milyar USD (÷1000)
        6. TOPLAM-STOK Alım/Satım Yönlü → Net Swap
        7. Altın TL Alım/Satım → Altın Net Swap

    Returns:
        DataFrame[index=Tarih, columns={
            'Tcmb Günlük Swap Net Alım (Milyar USD)',
            'Tcmb Günlük Altın Swap Net Alım (Milyar USD)',
            ... (tüm alt sütunlar)
        }]
    """
    from bs4 import BeautifulSoup
    import pdfplumber

    headers = {"User-Agent": _DEFAULT_USER_AGENT}

    # 1) Sayfa HTML + PDF link bul
    page_bytes = _safe_get(url, headers=headers, timeout=timeout,
                           max_bytes=10 * 1024 * 1024)
    html_text = page_bytes.decode("utf-8", errors="replace")

    soup = BeautifulSoup(html_text, "html.parser")

    def _normalize_tr(s: str) -> str:
        s = (s or "").lower()
        # Turkish capital İ lowercases to i + U+0307 (combining dot); strip it.
        s = s.replace("̇", "")
        for tr, en in [("ı", "i"), ("ş", "s"), ("ğ", "g"), ("ç", "c"), ("ö", "o"), ("ü", "u")]:
            s = s.replace(tr, en)
        return s

    # PDF linki bul: keyword match yeterli — TCMB CMS URL'lerinde .pdf uzantısı
    # olmayabilir (örn /wps/wcm/connect/a6ffdb2f-.../TCMB+Tarafli+Swap...).
    # Magic check (%PDF) zaten _safe_get sonrası yapılır.
    pdf_link = None
    for a in soup.find_all("a"):
        href = (a.get("href") or "").strip()
        title = (a.get("title") or "").strip()
        text_inner = a.get_text() or ""
        if not href:
            continue

        # Şüpheli olmayan linklere bak: ya .pdf uzantısı ya da gövde/title'da
        # PDF anahtar kelimelerinin tümü geçsin (TCMB CMS pattern)
        text_all = _normalize_tr(text_inner) + " " + _normalize_tr(title)
        keywords_match = all(k in text_all for k in PDF_KEYWORDS)
        if not keywords_match:
            continue
        pdf_link = href
        break

    if not pdf_link:
        raise RuntimeError(
            f"Taraflı Swap PDF linki bulunamadı (anahtar kelimeler: {PDF_KEYWORDS})"
        )

    if not pdf_link.lower().startswith("http"):
        pdf_link = "https://www.tcmb.gov.tr" + pdf_link

    # SSRF doğrulama: relatif birleştirme veya absolute href kontrolü
    _assert_allowed_host(pdf_link)

    # 2) PDF indir (boyut kapısı + stream)
    pdf_bytes = _safe_get(pdf_link, headers=headers, timeout=timeout,
                          max_bytes=_MAX_PDF_BYTES)
    if not pdf_bytes[:4] == b"%PDF":
        raise RuntimeError("İndirilen dosya PDF değil (magic check fail)")

    # 3) pdfplumber ile tüm satırları topla — header'ı şimdi ayırmıyoruz.
    # M kodunda olduğu gibi, tüm raw satırlar bir havuza alınır, sonra
    # "Valör/Value" pattern'i ile header satırı dinamik tespit edilir.
    all_rows: list[list] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for tbl in tables:
                if not tbl:
                    continue
                for row in tbl:
                    if row and any(c for c in row):
                        all_rows.append(list(row))

    if not all_rows:
        raise RuntimeError("PDF içinde tablo bulunamadı")

    # 4) En geniş tabloları al (max sütun sayısı)
    max_cols = max(len(r) for r in all_rows)
    if max_cols < 5:
        raise RuntimeError(
            f"PDF tablolarının max sütun sayısı {max_cols}, beklenen >=5."
        )
    wide_rows = [r for r in all_rows if len(r) == max_cols]

    # 5) Sütun temizleyici (newline/tab/dipnot)
    def _clean_col(c: Any) -> str:
        s = str(c) if c is not None else ""
        s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ").strip()
        if "{" in s:
            s = s.split("{", 1)[0].strip()
        s = " ".join(s.split())
        return s

    # 6) Header satırını bul — "Valör"/"Value" pattern'i ilk hücrede
    def _is_header_row(row: list) -> bool:
        if not row or row[0] is None:
            return False
        first = _normalize_tr(str(row[0]))
        return "valor" in first or "value" in first or "tarih" in first

    header_row_idx: int | None = None
    for i, row in enumerate(wide_rows):
        if _is_header_row(row):
            header_row_idx = i
            break

    if header_row_idx is None:
        # Fallback: ilk satırı header say (pdfplumber zaten ayırmış olabilir)
        header_row_idx = 0

    headers_raw = wide_rows[header_row_idx]
    data_rows = wide_rows[header_row_idx + 1 :]

    # 7) Sütun adlarını temizle ve dedupe et
    seen: dict[str, int] = {}
    column_names: list[str] = []
    for idx, h in enumerate(headers_raw):
        base = _clean_col(h) or f"col_{idx}"
        if base in seen:
            seen[base] += 1
            base = f"{base}__{seen[base]}"
        else:
            seen[base] = 0
        column_names.append(base)

    # 8) DataFrame oluştur — data satırlarını uniform uzunluğa getir
    ncols = len(column_names)
    normalized_data: list[list] = []
    for row in data_rows:
        if len(row) >= ncols:
            normalized_data.append(list(row[:ncols]))
        else:
            normalized_data.append(list(row) + [None] * (ncols - len(row)))
    combined = pd.DataFrame(normalized_data, columns=column_names)

    # 9) İlk sütunu Tarih olarak adlandır
    first_col_name = combined.columns[0]
    if first_col_name != "Tarih":
        combined = combined.rename(columns={first_col_name: "Tarih"})

    combined["Tarih"] = pd.to_datetime(
        combined["Tarih"], dayfirst=True, errors="coerce"
    )
    combined = combined[combined["Tarih"].notna()].reset_index(drop=True)

    # 5) Sayısal sütunları çevir + Milyon USD → Milyar USD (÷1000)
    for col in combined.columns:
        if col == "Tarih":
            continue
        combined[col] = combined[col].apply(_parse_pdf_number)
        combined[col] = combined[col] / 1000.0

    # 6) Net hesaplar
    alim_col = _find_col(combined, ["toplam", "stok", "alim", "yonlu"])
    satim_col = _find_col(combined, ["toplam", "stok", "satim", "yonlu"])

    if alim_col and satim_col:
        a = combined[alim_col].fillna(0)
        s = combined[satim_col].fillna(0)
        combined["Tcmb Günlük Swap Net Alım (Milyar USD)"] = a - s

    # Altın swap kalemleri
    altin_alim_col = _find_col(combined, ["altin", "tl ", "alim", "yonlu", "stok"])
    altin_satim_col = _find_col(combined, ["altin", "tl ", "satim", "yonlu", "stok"])

    if altin_alim_col and altin_satim_col:
        a = combined[altin_alim_col].fillna(0)
        s = combined[altin_satim_col].fillna(0)
        combined["Tcmb Günlük Altın Swap Net Alım (Milyar USD)"] = a - s

    # Döviz net alım kalemini de çıkartabilmek için (Toplam - Altın)
    if "Tcmb Günlük Swap Net Alım (Milyar USD)" in combined.columns and \
       "Tcmb Günlük Altın Swap Net Alım (Milyar USD)" in combined.columns:
        combined["Tcmb Günlük Döviz Swap Net Alım (Milyar USD)"] = (
            combined["Tcmb Günlük Swap Net Alım (Milyar USD)"]
            - combined["Tcmb Günlük Altın Swap Net Alım (Milyar USD)"]
        )

    # Tarih sıralı index'le döner
    combined = combined.sort_values("Tarih").set_index("Tarih")

    return combined


def _parse_pdf_number(v: Any) -> float | None:
    """PDF'teki sayıyı float'a çevir. TR/EN format toleranslı.

    '1,246.50' (EN) → 1246.5
    '1.246,50' (TR) → 1246.5
    '1.246'   → binlik (3 hane sonra) ya da 1.246?  → 3 hane sonra ise binlik (TR)
    '1.25'    → ondalık (1 hane)
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    if not s:
        return None
    # Boşluk ve nbsp kaldır
    s = s.replace(" ", "").replace("\xa0", "")
    # Sadece geçerli karakterler
    s = re.sub(r"[^0-9.,\-]", "", s)
    if not s:
        return None

    has_dot = "." in s
    has_comma = "," in s

    try:
        if has_dot and has_comma:
            # Önce gelen ayraç binlik
            dot_pos = s.rfind(".")
            comma_pos = s.rfind(",")
            if dot_pos < comma_pos:
                # TR: 1.246,50
                s2 = s.replace(".", "").replace(",", ".")
            else:
                # EN: 1,246.50
                s2 = s.replace(",", "")
            return float(s2)
        elif has_dot and not has_comma:
            # Sadece nokta
            dot_pos = s.rfind(".")
            after = len(s) - dot_pos - 1
            if after == 3:
                # binlik (TR), örn 1.246
                return float(s.replace(".", ""))
            else:
                # ondalık (EN)
                return float(s)
        elif has_comma and not has_dot:
            comma_pos = s.rfind(",")
            after = len(s) - comma_pos - 1
            if after == 3:
                # binlik (EN), örn 1,246
                return float(s.replace(",", ""))
            else:
                # ondalık (TR)
                return float(s.replace(",", "."))
        else:
            return float(s)
    except (ValueError, AttributeError):
        return None


def _find_col(df: pd.DataFrame, keywords: list[str]) -> str | None:
    """Sütunlardan tüm anahtar kelimeleri içereni bul (TR-normalize)."""
    def _normalize_tr(s: str) -> str:
        s = (s or "").lower()
        # Turkish capital İ lowercases to i + U+0307 (combining dot); strip it.
        s = s.replace("̇", "")
        for tr, en in [("ı", "i"), ("ş", "s"), ("ğ", "g"), ("ç", "c"), ("ö", "o"), ("ü", "u")]:
            s = s.replace(tr, en)
        return s

    keywords_n = [_normalize_tr(k) for k in keywords]
    for col in df.columns:
        col_n = _normalize_tr(str(col))
        if all(k in col_n for k in keywords_n):
            return str(col)
    return None
